"""Additional feature endpoints: EVM, Gantt, Burndown, Workload, Portfolio, Export."""

import io
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import String, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user
from app.models.project import Project
from app.models.task import Task
from app.models.task_dependency import TaskDependency
from app.models.team_member import TeamMember
from app.models.risk import Risk
from app.models.deliverable import Deliverable
from app.models.user import User
from app.services.evm import compute_evm
from app.services.gantt import build_gantt_data
from app.services.burndown import compute_burndown
from app.services.schedule import TaskNode, Dependency, compute_cpm

router = APIRouter(prefix="/api/projects", tags=["features"], dependencies=[Depends(get_current_user)])


# ── EVM ─────────────────────────────────────────────────────────────

@router.get("/{project_id}/evm")
async def get_evm(project_id: UUID, db: AsyncSession = Depends(get_db)):
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    tasks = (await db.execute(select(Task).where(Task.project_id == project_id))).scalars().all()
    task_data = [
        {"planned_cost": t.planned_cost, "actual_cost": t.actual_cost, "status": t.status.value}
        for t in tasks
    ]

    result = compute_evm(task_data, project.budget)
    return {
        "bac": result.bac, "pv": result.pv, "ev": result.ev, "ac": result.ac,
        "sv": result.sv, "cv": result.cv, "spi": result.spi, "cpi": result.cpi,
        "eac": result.eac, "etc": result.etc, "vac": result.vac, "tcpi": result.tcpi,
        "percent_complete": result.percent_complete, "percent_spent": result.percent_spent,
    }


# ── Gantt ───────────────────────────────────────────────────────────

@router.get("/{project_id}/gantt")
async def get_gantt(project_id: UUID, db: AsyncSession = Depends(get_db)):
    tasks = (await db.execute(select(Task).where(Task.project_id == project_id))).scalars().all()
    deps = (await db.execute(select(TaskDependency).where(TaskDependency.project_id == project_id))).scalars().all()

    nodes = [
        TaskNode(id=str(t.id), title=t.title, duration=t.duration_days or 0,
                 optimistic=t.optimistic_duration, most_likely=t.most_likely_duration,
                 pessimistic=t.pessimistic_duration, status=t.status.value)
        for t in tasks
    ]
    edges = [
        Dependency(predecessor_id=str(d.predecessor_id), successor_id=str(d.successor_id),
                   dep_type=d.dependency_type.value, lag=d.lag_days)
        for d in deps
    ]

    cpm_result = compute_cpm(nodes, edges)
    cpm_tasks = [
        {"id": t.id, "title": t.title, "duration": t.duration, "es": t.es, "ef": t.ef,
         "ls": t.ls, "lf": t.lf, "is_critical": t.is_critical, "status": t.status}
        for t in cpm_result.tasks
    ]

    # Get assignee names
    member_map: dict[str, str] = {}
    for t in tasks:
        if t.assignee_id:
            member_map[str(t.id)] = str(t.assignee_id)
    members = (await db.execute(select(TeamMember))).scalars().all()
    name_map = {str(m.id): m.name for m in members}

    raw = [
        {"id": str(t.id), "wbs_code": t.wbs_code, "parent_id": str(t.parent_id) if t.parent_id else None,
         "is_milestone": t.is_milestone, "status": t.status.value,
         "assignee_name": name_map.get(member_map.get(str(t.id), ""))}
        for t in tasks
    ]
    dep_data = [
        {"predecessor_id": str(d.predecessor_id), "successor_id": str(d.successor_id)}
        for d in deps
    ]

    bars = build_gantt_data(cpm_tasks, dep_data, raw)
    return {
        "project_duration": cpm_result.project_duration,
        "bars": [
            {"id": b.id, "title": b.title, "wbs_code": b.wbs_code, "parent_id": b.parent_id,
             "start_day": b.start_day, "end_day": b.end_day, "duration": b.duration,
             "status": b.status, "is_critical": b.is_critical, "is_milestone": b.is_milestone,
             "assignee": b.assignee, "progress": b.progress, "dependencies": b.dependencies}
            for b in bars
        ],
    }


# ── Burndown ────────────────────────────────────────────────────────

@router.get("/{project_id}/burndown")
async def get_burndown(project_id: UUID, db: AsyncSession = Depends(get_db)):
    tasks = (await db.execute(select(Task).where(Task.project_id == project_id))).scalars().all()
    task_data = [
        {"story_points": t.story_points, "created_at": t.created_at.isoformat() if t.created_at else None,
         "completed_date": t.completed_date.isoformat() if t.completed_date else None}
        for t in tasks
    ]
    return compute_burndown(task_data)


# ── Workload ────────────────────────────────────────────────────────

@router.get("/{project_id}/workload")
async def get_workload(project_id: UUID, db: AsyncSession = Depends(get_db)):
    members = (await db.execute(
        select(TeamMember).where(TeamMember.project_id == project_id)
    )).scalars().all()

    result = []
    for m in members:
        tasks = (await db.execute(
            select(Task).where(Task.project_id == project_id, Task.assignee_id == m.id)
        )).scalars().all()

        total_tasks = len(tasks)
        active = sum(1 for t in tasks if t.status.value in ("in_progress", "in_review"))
        done = sum(1 for t in tasks if t.status.value == "done")
        total_hours = sum(((t.duration_days or 0) * 8) for t in tasks if t.status.value != "done")

        result.append({
            "id": str(m.id), "name": m.name, "role": m.role.value,
            "availability": m.availability,
            "total_tasks": total_tasks, "active_tasks": active, "done_tasks": done,
            "remaining_hours": round(total_hours, 1),
            "utilization": round(total_hours / (m.availability * 8 / 100) * 100, 1) if m.availability and m.availability > 0 else 0,
        })

    result.sort(key=lambda x: x["utilization"], reverse=True)
    return result


# ── Portfolio ───────────────────────────────────────────────────────

@router.get("/portfolio/overview", tags=["portfolio"])
async def get_portfolio(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cross-project dashboard."""
    projects = (await db.execute(select(Project).order_by(Project.created_at.desc()))).scalars().all()

    portfolio = []
    for p in projects:
        task_count = await db.scalar(select(func.count(Task.id)).where(Task.project_id == p.id)) or 0
        done_count = await db.scalar(
            select(func.count(Task.id)).where(Task.project_id == p.id, cast(Task.status, String) == "DONE")
        ) or 0
        risk_count = await db.scalar(
            select(func.count(Risk.id)).where(
                Risk.project_id == p.id,
                cast(Risk.status, String).notin_(["RESOLVED", "CLOSED"]),
            )
        ) or 0

        portfolio.append({
            "id": str(p.id), "name": p.name, "status": p.status.value,
            "approach": p.development_approach.value, "budget": p.budget,
            "total_tasks": task_count, "done_tasks": done_count,
            "completion_pct": round(done_count / task_count * 100, 1) if task_count else 0,
            "open_risks": risk_count,
        })

    return portfolio


# ── Export Excel ────────────────────────────────────────────────────

@router.get("/{project_id}/export/excel")
async def export_excel(project_id: UUID, db: AsyncSession = Depends(get_db)):
    from openpyxl import Workbook

    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    wb = Workbook()

    # Tasks sheet
    ws = wb.active
    ws.title = "Tasks"
    ws.append(["Title", "Status", "Priority", "Duration", "Planned Cost", "Actual Cost", "Story Points", "WBS", "Milestone"])
    tasks = (await db.execute(select(Task).where(Task.project_id == project_id))).scalars().all()
    for t in tasks:
        ws.append([t.title, t.status.value, t.priority.value, t.duration_days, t.planned_cost, t.actual_cost, t.story_points, t.wbs_code, t.is_milestone])

    # Risks sheet
    ws2 = wb.create_sheet("Risks")
    ws2.append(["Title", "Category", "Probability", "Impact", "Status", "Strategy", "Response Plan"])
    risks = (await db.execute(select(Risk).where(Risk.project_id == project_id))).scalars().all()
    for r in risks:
        ws2.append([r.title, r.category.value, r.probability.value, r.impact.value, r.status.value, r.strategy.value, r.response_plan])

    # Deliverables sheet
    ws3 = wb.create_sheet("Deliverables")
    ws3.append(["Name", "Status", "Quality", "Completion %", "Acceptance Criteria"])
    deliverables = (await db.execute(select(Deliverable).where(Deliverable.project_id == project_id))).scalars().all()
    for d in deliverables:
        ws3.append([d.name, d.status.value, d.quality_level.value, d.completion_percentage, d.acceptance_criteria])

    # Team sheet
    ws4 = wb.create_sheet("Team")
    ws4.append(["Name", "Role", "Email", "Availability", "Skills"])
    members = (await db.execute(select(TeamMember).where(TeamMember.project_id == project_id))).scalars().all()
    for m in members:
        ws4.append([m.name, m.role.value, m.email, m.availability, m.skills])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={project.name.replace(' ', '_')}_export.xlsx"},
    )
